#import openpyxl
#wb = openpyxl.load_workbook('//home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
#print(type(wb))

#import openpyxl
#wb = openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLS_10.xls')
#print(wb.sheetnames)
#sheet = wb['Sheet1']
#print(sheet)
#print(type(sheet))
#print(sheet.title)
#anotherSheet = wb.active
#print(anotherSheet)


#import openpyxl
#wb = openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
#sheet = wb['Sheet1']
#c= sheet['C3'].value
#print(c)

#import openpyxl
#wb= openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
#sheet = wb['Sheet1']
#print(sheet['A1'])
#print(sheet['A1'].value)
#c = sheet['B1']
#print(c.value)
#print('Row %s, Column %s is %s' % (c.row, c. column, c.value))
#print(sheet.max_row)
#print(sheet.max_column)



#import openpyxl
#wb = openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
#sheet = wb['Sheet1']
#for i in range(1,8,2):
#    print(i, sheet.cell(row=i, column=2).value)


#import openpyxl
#from openpyxl.utils import get_column_letter, column_index_from_string
#print(get_column_letter(900))
#wb = openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
#sheet = wb['Sheet1']
#print('get_column_letter(sheet.max_column)')

import openpyxl
wb = openpyxl.load_workbook('/home/ashok/Documents/debugging/file_example_XLSX_10.xlsx')
sheet = wb['Sheet1']
print(tuple(sheet['A1':'C3']))